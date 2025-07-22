'''
This is a GUI program desined to pull CAISO OASIS LMP Pricing data, 
cleaned and organized into an excel spreadsheet. 
'''

import os
import sys
import time
from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile
import tkinter as tk
from tkinter import filedialog

import pandas as pd
import requests
import matplotlib.pyplot as plt
import seaborn as sns

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

from pydantic import ValidationError

from tkcalendar import Calendar
from customtkinter import CTk, CTkButton, CTkLabel, CTkComboBox, CTkEntry, set_appearance_mode

# Market configuration- centralized settings
MARKET_CONFIG = {
    'DAM': {
        'queryname': 'PRC_LMP',
        'version': 1,
        'interval_minutes': 60,
        'value_column': 'MW',
        'has_greenhouse_gas': True,
        'report_name': 'Locational Marginal Prices'
    },
    'RTM': {
        'queryname': 'PRC_INTVL_LMP',
        'version': 3,
        'interval_minutes': 5,
        'value_column': 'VALUE',
        'has_greenhouse_gas': True,
        'report_name': 'Interval Locational Marginal Prices'
    },
    'HASP': {
        'queryname': 'PRC_HASP_LMP',
        'version': 1,
        'interval_minutes': 15,
        'value_column': 'MW',
        'has_greenhouse_gas': False,
        'report_name': 'Hour Ahead Locational Marginal Prices'
    },
    'FMM': {
        'queryname': 'PRC_RTPD_LMP',
        'version': 2,
        'interval_minutes': 15,
        'value_column': 'PRC',
        'has_greenhouse_gas': True,
        'report_name': 'FMM Locational Marginal Prices'
    }
}

# Configuration to get related variables based on market run id
def get_market_config(market_run_id):
    '''
    This method calls the MARKET_CONFIG map and links it to the
     appropriate variables based on market_run_id
    '''
    return MARKET_CONFIG.get(market_run_id, MARKET_CONFIG['DAM'])

# Pulls, cleans, and formats data into the excel file.
def backend(market_run_id, startdate, enddate):
    '''
    This method runs everything the GUI does after the submit button
    is pressed. It calls the API, pulls the cleaned data into an
    excel file, and adds 3 analysis sheets: hourly average, monthly
    average, and summary statistics
    '''
    node = node_var.get()
    node = node.replace(' ', '')
    # API Calling Function calls each 30 day chunk and adds it to a list of files
    def pull_request(startdate, enddate):
        startdate = int(startdate.strftime('%Y%m%d'))  # Updating the date format
        enddate = int(enddate.strftime('%Y%m%d'))


        url = "http://oasis.caiso.com/oasisapi/SingleZip"
        params = {
            "resultformat": 6,  # Resultformat should always be 6- it creates a CSV.
            "queryname": queryname,  # Represents the report name, based on market_run_id
            "startdatetime": f'{startdate}T07:00-0000',  # Dates start at 7 hour for MST purposes
            "enddatetime": f'{enddate}T07:00-0000', 
            "market_run_id": market_run_id,
            "version": version,  
            "node": node,
            }

        response = requests.get(url, params=params, timeout=1000)
        try:  # Try reading the file (or each chunk), looking for query/parameter issues
            with ZipFile(BytesIO(response.content)) as z:
                for filename in z.namelist():
                    with z.open(filename) as f:
                        df = pd.read_csv(f)
                        df.to_csv(f'pull#{counter}.csv')
                        if '<?xml version="1.0" encoding="UTF-8"?>' in df.columns:
                            pass  # Not adding in files that are empty
                        else:  # Appends files that pull correctly
                            files.append(f'pull#{counter}.csv')  # Listing files to combine
        except ValidationError as e:  # Catching errors with API Input
            print(f'Error with Params: {e}')
            status_lbl.configure(text='Error with Inputs. Please ' \
            'check dates, node(s), and output folder.')
            root.update()
        except Exception as e:  # Catch all, usually not errors
            print(f"Error: {e}")

    # Cleaning each 30 day file
    def clean_file(filename):
        '''
        This method cleans the first/main report in the spreadsheet,
        includeing dropping and renaming columns, and formatting it
        for readability
        '''
        df = pd.read_csv(filename)
        all_drop = ['NODE_ID_XML', 'NODE_ID', 'PNODE_RESMRID',  # Dropping unneeded columns
                    'OPR_DT', 'OPR_HR', 'OPR_INTERVAL',
                    'XML_DATA_ITEM', 'POS', 'GROUP',
                    'GRP_TYPE', 'MARKET_RUN_ID', 'Unnamed: 0',
                    'INTERVAL_START_TIME']
        valid_columns = [col for col in all_drop if col in df.columns]
        df = df.drop(columns=valid_columns)  # Dropping conditinally
        df = df.sort_values(['INTERVALSTARTTIME_GMT'])

        # Changing the timezone to MST
        df['INTERVALSTARTTIME_GMT'] = df['INTERVALSTARTTIME_GMT'].str.replace('-00:00', '')
        df['INTERVALSTARTTIME_GMT'] = df['INTERVALSTARTTIME_GMT'].str.replace('T', ' ')
        df['INTERVALENDTIME_GMT'] = df['INTERVALENDTIME_GMT'].str.replace('-00:00','')
        df['INTERVALENDTIME_GMT'] = df['INTERVALENDTIME_GMT'].str.replace('T',' ')
        # Shifting the hours back by 7 to align with MST
        df['INTERVALSTARTTIME_GMT'] = pd.to_datetime(df['INTERVALSTARTTIME_GMT'])
        df['INTERVALSTARTTIME_GMT'] = df['INTERVALSTARTTIME_GMT'] - timedelta(hours=7)
        df['INTERVALENDTIME_GMT'] = pd.to_datetime(df['INTERVALENDTIME_GMT'])
        df['INTERVALENDTIME_GMT'] = df['INTERVALENDTIME_GMT'] - timedelta(hours=7)
        df.rename(columns={'INTERVALSTARTTIME_GMT': 'INTERVALSTARTTIME_MST',
                           'INTERVALENDTIME_GMT': 'INTERVALENDTIME_MST'}, inplace=True)

        # Conditional cleaning based on column name for price
        if 'LMP_TYPE' in df.columns:
            df['LMP_TYPE'] = df['LMP_TYPE'].replace({'LMP':'LMP', 'MCC':'Congestion',
                                                     'MCE':'Energy', 'MCL':'Loss',
                                                     'MGHG':'Greenhouse Gas'})

        # Splitting date into smaller columns for readability and grouping
        df['INTERVALSTARTTIME_MST'] = df['INTERVALSTARTTIME_MST'].astype(str)
        df[['Year', 'Month', 'Day']] = df['INTERVALSTARTTIME_MST'].str.split('-',expand=True)
        df[['Day', 'Time']] = df['Day'].str.split(' ',expand=True)
        df[['Hour (MST)', 'Minute', 'Seconds']] = df['Time'].str.split(':',expand=True)
        df = df.drop(columns=['Time', 'Seconds'])
        df.to_csv(filename)  # Replacing file with cleaned version

    # Finding interval rows that are missing from the DF and adding in the missing values
    def fill_missing_values(filename, market_run_id):
        '''
        This method finds all of the missing intervals not found in the
        df. It fills in the missing values and backfills LMP values.
        '''
        config = get_market_config(market_run_id)
        interval_minutes = config['interval_minutes']

        df = pd.read_excel(filename)
        # Finding and creating rows for missing intervals
        dt = pd.to_datetime(df['INTERVALSTARTTIME_MST'])
        df['INTERVALSTARTTIME_MST'] = dt.dt.floor(f'{interval_minutes}min')
        # Finding all intervals I should have
        full_range = pd.date_range(start=df['INTERVALSTARTTIME_MST'].min(),
                                   end=df['INTERVALSTARTTIME_MST'].max(),
                                   freq=f'{interval_minutes}min')
        full_df = pd.DataFrame({'INTERVALSTARTTIME_MST': full_range})  # New df for all intervals
        if 'Greenhouse Gas' in full_df.columns:  # Conditional logic for HASP
            result = full_df.merge(
                df[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                    'NODE', 'Year', 'Month', 'Hour (MST)', 'Minute',
                    'Congestion', 'Energy', 'Loss', 'Greenhouse Gas',
                    'LMP']],
                on='INTERVALSTARTTIME_MST',
                how='outer'
            )
        else:
            result = full_df.merge(
                df[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                    'NODE', 'Year', 'Month', 'Hour (MST)', 'Minute',
                    'Congestion', 'Energy', 'Loss', 'LMP']],
                on='INTERVALSTARTTIME_MST',
                how='outer'
            )
        # Filling in missing values
        interval_end = result['INTERVALSTARTTIME_MST'] + pd.Timedelta(minutes=interval_minutes)
        result['INTERVALENDTIME_MST'] = result['INTERVALENDTIME_MST'].fillna(interval_end)
        result = result.sort_values(['INTERVALSTARTTIME_MST', 'NODE'])  # Sort by node to backfill
        result['NODE'] = result['NODE'].bfill()
        result['INTERVALSTARTTIME_MST'] = result['INTERVALSTARTTIME_MST'].astype(str)
        if result['INTERVALENDTIME_MST'].isnull:  # If it has missing values, fill them in!
            result[['Year', 'Month', 'Day']] = result['INTERVALSTARTTIME_MST'].str.split('-',
                expand=True)
            result[['Day', 'Time']] = result['Day'].str.split(' ',expand=True)
            result[['Hour (MST)', 'Minute', 'Seconds']] = result['Time'].str.split(':',expand=True)
        result = result.drop(columns=['Time', 'Seconds']).sort_values(['Hour (MST)', 'Minute'])

        # Filling in LMP values, backfilling with previous day's same hour and minute intvl price.
        if 'Greenhouse Gas' in result.columns:
            LMP_list = ['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']
        else:
            LMP_list = ['Congestion', 'Energy', 'Loss', 'LMP']
        for column in LMP_list:
            result[column] = result[column].bfill()
        result = result.sort_values('INTERVALSTARTTIME_MST')  # Setting to original sorting
        # Reordering columns
        if 'Greenhouse Gas' in result.columns:
            result = result[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                             'NODE', 'Year', 'Month', 'Day', 'Hour (MST)',
                             'Minute', 'LMP','Congestion', 'Energy', 'Loss',
                             'Greenhouse Gas']]
        else:
            result = result[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                             'NODE', 'Year', 'Month', 'Day', 'Hour (MST)',
                             'Minute', 'LMP','Congestion', 'Energy', 'Loss']]
        result.to_excel(filename)

    # Creating the monthly average sheet and chart
    def monthly_average(filename):
        df = pd.read_excel(filename)
        # Create a new column containing month and year (spelled out) and have this be the X axis
        df['Date'] = df['Month'].astype(str).str.zfill(2) + '/01/' + df['Year'].astype(str)
        df['Date'] = pd.to_datetime(df['Date'], format='%m/%d/%Y')
        df['Date'] = df['Date'].dt.strftime('%B %Y')  # Prints date spelled out (May 2024)
        df = df.sort_values(['Year', 'Month']) # sorts first by year, then by month

        if 'Greenhouse Gas' in df.columns:
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Date'],
                                as_index=False)[['Congestion', 'Energy',
                                                 'Greenhouse Gas', 'LMP', 'Loss']].mean() 
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Date', 'LMP',
                             'Congestion', 'Energy', 'Greenhouse Gas', 'Loss']]
        else:
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Date'],
                                as_index=False)[['Congestion', 'Energy',
                                                 'LMP', 'Loss']].mean()
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Date', 'LMP',
                             'Congestion', 'Energy', 'Loss']]

        # Adding sheet to excel file
        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx',
                            engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_avg.to_excel(writer, sheet_name='Monthly Average', index=False)
        plt.figure(figsize=(10,8))
        plt.plot(df_avg['Date'], df_avg['LMP'])  # Creating monthly average line chart
        plt.tick_params(axis='x', labelrotation=45)
        plt.grid()
        plt.title(f'Monthly Average LMP {node}')
        plt.ylabel('Avg $/MWh')
        plt.xlabel('Month')
        plt.savefig(f'{output_file_path}/monthlyline.png')
        plt.close()  # Closing plt so it doesn't combine with other chart later
        img_path = f'{output_file_path}/monthlyline.png'
        add_chart_to_excel(filename, 'Monthly Average', img_path, 'J1')

        format_excel_cells('Monthly Average', 4, 10)  # Calling number formatting function

    # Creating hourly average sheet and heatmap
    def hourly_average(filename):
        '''
        This method takes the hourly average of the main report and
        adds it to a new sheet. It also includes a heatmap for the
        12X24 data, averaging multiple years.
        '''
        df = pd.read_excel(filename)
        if 'Greenhouse Gas' in df.columns:
            df_avg = df.groupby(['NODE', 'Date', 'Day', 'Hour (MST)'],
                                as_index=False)[['Congestion', 'Energy', 'Greenhouse Gas', 'LMP',
                                                 'Loss']].mean()
            df_avg[['Congestion', 'Energy', 'Loss', 'Greenhouse Gas',
                    'LMP']] = df_avg[['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']]
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Congestion', 'Energy',
                             'Greenhouse Gas', 'Loss', 'LMP']]  # Reordering column names
        else:
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Day', 'Hour (MST)'],
                                as_index=False)[['Congestion', 'Energy', 'LMP', 'Loss']].mean()
            df_avg[['Congestion', 'Energy', 'Loss','LMP']] = df_avg[['Congestion', 'Energy',
                                                                     'Loss', 'LMP']]
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Congestion', 'Energy',
                             'Loss', 'LMP']]

        count = (df_avg['LMP'] < 0).sum()  # Counting how many hours LMP is below 0

        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx',
                            engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Witing hourly averages sheet
            df_avg.to_excel(writer, sheet_name='Hourly Average', index=False)
            row = 13  # Adding blank rows in between
            pd.DataFrame([['Number of hours LMP is below 0:']]).to_excel(writer,
                sheet_name='Summary Statistics', startrow=row, header=False, index=False)
            row += 1
            pd.DataFrame([[count]]).to_excel(writer, sheet_name='Summary Statistics',
                                             startrow=row, header=False, index=False)
            row += 2
            pd.DataFrame([['Duration Curve']]).to_excel(writer, sheet_name='Summary Statistics',
                                                        startrow=row, header=False, index=False)
            row += 1

        # Heat map logic
        pivot = df.pivot_table(index='Hour (MST)', columns='Month', values='LMP', aggfunc='mean')
        plt.figure(figsize=(10,8))
        sns.heatmap(pivot, annot=True, cmap='RdYlGn_r', fmt='.0f', cbar_kws={'label':'$/MWH'})
        plt.title(f'12x24 Heatmap {node}')
        plt.savefig(f'{output_file_path}/heatmap.png')
        plt.close()
        img_path = f'{output_file_path}/heatmap.png'
        add_chart_to_excel(filename, 'Hourly Average', img_path, 'L3')  # Adding chart to excel
        format_excel_cells('Hourly Average', 6,10)  # Calling number formatting function

    # Creating summary statistics sheet
    def summary_statistics(filename):
        '''
        This method creates a new sheet depicting summary stats for
        the four LMP values.
        '''
        df = pd.read_excel(filename)
        if 'Greenhouse Gas' in df.columns:
            cols = ['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']
        else:
            cols = ['Congestion', 'Energy', 'Loss', 'LMP']
        desc = df[cols].describe()  # Finding what I want to display
        desc = desc.round(4)

        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx',
                            engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            row = 0
            pd.DataFrame([['Summary Statistics']]).to_excel(writer,
                                                            sheet_name='Summary Statistics',
                                                            startrow=row, header=False,
                                                            index=False)
            row += 1
            desc.to_excel(writer, sheet_name = 'Summary Statistics', startrow=row, header=True,
                          index=True)
            row += 13

        format_excel_cells('Summary Statistics', 2, 5)  # Calling number formatting function

    # Creating duration charts and adding them to the sheet
    def duration_chart(filename):
        '''
        This method creates a duration chart for the entire report, as
        well as two zoomed charts, for the first and last 5%.
        '''
        # Cleaning to get chart columns
        df = pd.read_excel(filename).sort_values('LMP', ascending=False)
        duration_counts = df['LMP'].value_counts()
        total_count = df['LMP'].value_counts().sum()
        df['duration_count'] = df['LMP'].map(duration_counts)  # Represents count of that value
        df['percent'] = df['duration_count']/total_count  # % of how often value appears in the df
        chart_lmp = df[['LMP', 'percent']].drop_duplicates().copy()
        chart_lmp['xval'] = chart_lmp['percent'].cumsum()
        xval_map = dict(zip(chart_lmp['LMP'], chart_lmp['xval']))
        df['xval'] = df['LMP'].map(xval_map)
        df = df[['LMP', 'xval']]

        # Writing the new df to a sheet and hiding it
        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx',
                            engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name='Hidden Duration Chart Data', index=False)
        wb = openpyxl.load_workbook(filename)
        ws = wb['Hidden Duration Chart Data']
        ws.sheet_state='hidden'  # Hiding the sheet

        # Page formatting, bolding titles
        sheet = wb['Summary Statistics']
        sheet['A1'].font = Font(bold=True)
        sheet['A14'].font = Font(bold=True)
        sheet['A17'].font = Font(bold=True)
        sheet['M17'].font = Font(bold=True)

        wb.save(filename)

        # Creating duration chart
        plt.figure()
        plt.scatter(df['xval'], df['LMP'], s=3)
        plt.axhline(y=0, color ='black')  # Creating 0 axis line
        plt.title(f'Duration Chart {node}')
        plt.ylabel('$/MWh')
        plt.xlabel('% of Time')
        plt.grid()
        plt.savefig(f'{output_file_path}/durationchart.png')  # Creating an image of the chart
        plt.close()
        img_path2 = f'{output_file_path}/durationchart.png'
        add_chart_to_excel(filename, 'Summary Statistics', img_path2, 'B19')

        # Creating lowest 5% zoom
        last5 = df.tail(int(len(df)*.05))
        plt.figure()
        plt.scatter(last5['xval'], last5['LMP'], s=3)
        plt.axhline(y=0, color='black')
        plt.title(f'Lowest 5% Zoom {node}')
        plt.ylabel('$/MWh')
        plt.xlabel('% of Time')
        plt.grid()
        plt.savefig(f'{output_file_path}/lowest5zoom.png')
        plt.close()
        img_path = f'{output_file_path}/lowest5zoom.png'
        add_chart_to_excel(filename, 'Summary Statistics', img_path, 'L46')

        # Creating highest 5% zoom
        first5 = df.head(int(len(df)*.05))
        plt.figure()
        plt.scatter(first5['xval'], first5['LMP'], s=3)
        plt.axhline(y=0, color='black')
        plt.title(f'Highest 5% Zoom {node}')
        plt.ylabel('$/MWh')
        plt.xlabel('% of Time')
        plt.grid()
        plt.savefig(f'{output_file_path}/highest5zoom.png')
        plt.close()
        img_path3 = f'{output_file_path}/highest5zoom.png'
        add_chart_to_excel(filename, 'Summary Statistics', img_path3, 'B46')

    # Formats all numbers in called rows to have two decimal places
    def format_excel_cells(sheet, min_col, max_col):
        '''
        This method formats the numerical values to two decimal places.
        '''
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet]
        for row in sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=2):
            for cell in row:
                cell.number_format = '0.00'
        wb.save(file)

    # Adding created chart to excel file
    def add_chart_to_excel(filename, sheet_name, chart_path, cell_reference):
        '''
        This method saves the python chart to an image and adds it to
        the excel sheet.
        '''
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        img = XLImage(chart_path)
        sheet.add_image(img, cell_reference)
        wb.save(filename)
        os.remove(chart_path)

    # Returning the appropriate LMP columns based on market_run_id
    def get_lmp_columns(has_greenhouse_gas=True):
        '''
        This method creates column lists based on whether
        'Greenhouse Gas' is a column.
        '''
        base_columns = ['Congestion', 'Energy', 'Loss', 'LMP']
        if has_greenhouse_gas:
            base_columns.insert(-1, 'Greenhouse Gas')  # Inserting before LMP
        return base_columns

    # Filtering columns depending on if it has greenhouse gas or not
    def get_ordered_columns(has_greenhouse_gas=True):
        base_columns = ['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST',
                        'NODE', 'Year', 'Month', 'Day', 'Hour (MST)',
                        'Minute']
        lmp_columns = get_lmp_columns(has_greenhouse_gas)
        return base_columns + lmp_columns


    # Preparing variables from user input
    get_market_config(market_run_id)
    config = get_market_config(market_run_id)
    queryname = config['queryname']
    version = config['version']
    startdate = datetime.strptime(startdate, '%m/%d/%y').date()
    enddate = datetime.strptime(enddate, '%m/%d/%y').date()
    enddate = enddate + timedelta(days=1)  # Adding on a day to pull so it gets the full last day
    difference = enddate - startdate
    days = difference.days  # Making a counter for my loop bc .days is readonly

    # Logic for using the API, cleaning, and formatting as an xlsx file.
    files = []
    df_list = []
    counter = 0
    timestamp = datetime.now()  # Using for filename
    timestamp = timestamp.strftime('%m-%d-%Y %H%M')

    while days > 0:  # Until we reach the end date
        if days < 30:  # When we get below 30 days left
            print(f'Pulling data for {startdate} to {enddate}.')
            pull_request(startdate, enddate)
            break  # Ending loop once I get to the end date
        nextdate = startdate + timedelta(days=30)  # Creating a chunk
        print(f'Pulling data for {startdate} to {nextdate}.')
        pull_request(startdate, nextdate)
        days -= 30
        counter += 1
        startdate = nextdate
        time.sleep(10)  # Waiting 10 seconds to avoid query limits
    for file in files:
        clean_file(file)
        df_list.append(pd.read_csv(file))  # Adding rows in the csv to df_list for concatenation

    # Combining files and cleaning them up a little
    df_combined = pd.concat(df_list, ignore_index=True)  # Combining 30 day chunks
    cond_drop = ['Unnamed: 0.1', 'Unnamed: 0']
    conditional_drop = [col for col in cond_drop if col in df_combined.columns]
    df_combined = df_combined.drop(columns=conditional_drop)
    df_combined = df_combined.drop_duplicates()  # Dropping duplicates

    # Pivoting table and reordering columns (for first sheet)
    get_market_config(market_run_id)
    config = get_market_config(market_run_id)
    value_column = config['value_column']
    has_greenhouse_gas = config['has_greenhouse_gas']

    # Breaking out LMP_TYPE columns, keeping the other indexed columns
    if value_column in df_combined.columns:
        df_combined = pd.pivot_table(df_combined, values=value_column,
                                     index=['INTERVALSTARTTIME_MST',
                                            'INTERVALENDTIME_MST',
                                            'NODE', 'Year', 'Month',
                                            'Day', 'Hour (MST)',
                                            'Minute'], columns='LMP_TYPE').reset_index()
        df_combined = df_combined[get_ordered_columns(has_greenhouse_gas)]

    # Pushing to excel file, deleting csv chunks
    file = f'{output_file_path}/{market_run_id} {timestamp}.xlsx'
    with pd.ExcelWriter(file, engine='openpyxl') as writer:
        df_combined.to_excel(writer, sheet_name='Report', index=False)
    fill_missing_values(file, market_run_id)
    monthly_average(file)
    hourly_average(file)
    summary_statistics(file)
    duration_chart(file)
    format_excel_cells('Sheet1', 10, 14)  # Formatting numbers on first sheet
    wb = openpyxl.load_workbook(file)
    sheet = wb['Sheet1']
    sheet.title='Report'  # Renaming the first sheet
    wb.save(file)

    # Deleting all of the csv 30 day chunk files from the folder
    if getattr(sys, 'frozen', False):  # Finding file path to wherever GUI is stored
        application_path = os.path.dirname(sys.executable)
    for csv in files:  # Deleting extra files
        filepath = os.path.join(application_path, csv)
        os.remove(filepath)

    status_lbl.configure(text='Finished!')  # Updating status label
    root.update()


# Functions for all the button/widgets of GUI
def submit():
    '''
    After user gives all inputs, this method runs all of the backend code.
    '''
    status_lbl.configure(text='Running...')
    root.update()  # Updating status label
    market_run_id=MRIDDropdown.get()  # Grabbing market_run_id based on user input
    backend(market_run_id, startdate, enddate)  # Calling backend code

def find_start_date():
    '''
    Creates start date based on user input
    '''
    global startdate
    startdate = cal.get_date()
    startdate_label.configure(text=f'Start date: {startdate}')

def find_end_date():
    '''
    Creates start date based on user input
    '''
    global enddate
    enddate = cal.get_date()
    enddate_label.configure(text=f'End date: {enddate}')

def select_output_file():
    '''
    Allows user to select an output location and creates a path
    '''
    global output_file_path
    directory = filedialog.askdirectory(title='Select output directory')
    if directory:
        output_file_path = directory
        output_file_label.configure(text=directory)  # Displaying file path
    else:
        output_file_label.configure(text='No directory selected yet')  # If submitted w/o filepath

def update_report_lbl(choice):
    '''
    Displays the report name based on the user chosen market type
    '''
    config = get_market_config(choice)
    report_name = config.get('report_name', 'Unknown Report')

    report_lbl.configure(text=f'{report_name}')
    root.update()


# Tkinter program
root = CTk()  # Initializing window
root.geometry('800x600')
set_appearance_mode('light')

node_var = tk.StringVar()

startdate = None  # Initializing
enddate = None

# Widgets
MRID_label = CTkLabel(root, text='Market Type:', font=('Arial',15), text_color='#04033A')
MRIDDropdown = CTkComboBox(master=root, values=['DAM', 'HASP','RTM', 'FMM'],
                           command=update_report_lbl)

report_lbl = CTkLabel(root, text='Locational Marginal Prices', font=('Arial', 15),
                      text_color='#04033A')

cal = Calendar(root, selectmode ='day',
            year=2024, month=1,  # Default setting
            day=1, font=('Arial', 15))

chooseStartDate = CTkButton(root, text='Choose Start Date', command=find_start_date,
                            corner_radius=26, fg_color='#162157', hover_color='#6D7DCF')
chooseEndDate = CTkButton(root, text='Choose End Date', command=find_end_date, corner_radius=26,
                          fg_color='#162157', hover_color='#6D7DCF')

startdate_label = CTkLabel(root, text= 'Start Date: ', font=('Arial', 15), text_color='#04033A')
enddate_label = CTkLabel(root, text='End Date: ', font=('Arial', 15), text_color='#04033A')

node_label = CTkLabel(root, text='Node(s):', font=('Arial', 15), text_color='#04033A')
node_entry = CTkEntry(root, textvariable = node_var, font=('Arial', 15), text_color='#04033A')

sub_btn=CTkButton(master=root,text = 'Submit', command=submit, corner_radius=32,
                  fg_color='#162157', hover_color='#6D7DCF')

output_file_button = CTkButton(root, text='Select Output File Path', command=select_output_file,
                               corner_radius=32, fg_color='#162157', hover_color='#6D7DCF')
output_file_label = CTkLabel(root, text='No path selected', font=('Arial', 10),
                             text_color='#04033A')

status_lbl = CTkLabel(root, text='', font=('Arial', 15), text_color='#04033A')

title_lbl = CTkLabel(root, text='CAISO OASIS DATA', font=('Arial', 20, 'bold'),
                     text_color='#04033A')

# Grid: where all the widgets are displayed on the GUI
cal.grid(row=6, column=0)
chooseStartDate.grid(row=4, column=0)
chooseEndDate.grid(row=5, column=0)
MRID_label.grid(row=1, column=1)
MRIDDropdown.grid(row=1, column=2)
node_label.grid(row=3, column=1)
node_entry.grid(row=3, column=2)
sub_btn.grid(row=6, column=2)
startdate_label.grid(row=4, column=1)
enddate_label.grid(row=5, column=1)
output_file_button.grid(row=1, column=0)
output_file_label.grid(row=2, column=0)
title_lbl.grid(row=0, column=2)
report_lbl.grid(row=2, column=2)
status_lbl.grid(row=7, column=2)

root.mainloop()  # Performing an infinite loop for the window to display
