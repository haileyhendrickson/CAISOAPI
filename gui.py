import tkinter as tk
from tkinter import filedialog, ttk
from tkinter import *
from tkcalendar import Calendar
from customtkinter import *
import pandas as pd 
import requests
from zipfile import ZipFile
from io import BytesIO
from datetime import timedelta, time, datetime
import openpyxl
import time
import sys
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import seaborn as sns


# all backend code
def backend(market_run_id, startdate, enddate): # Pulls, cleans, and formats data into the excel file.
    node=node_var.get()
    node = node.replace(' ', '')
    # API Calling Function
    def pull_request(startdate, enddate): # calls each 30 day chunk and adds it to a list of files
        startdate = int(startdate.strftime('%Y%m%d')) # updating date format
        enddate = int(enddate.strftime('%Y%m%d'))
        url = "http://oasis.caiso.com/oasisapi/SingleZip"
        params = {
            "resultformat": 6, # should always be this- creates a CSV
            "queryname": queryname, # locational marginal prices
            "startdatetime": f'{startdate}T07:00-0000',  # starting at the 7 hour for MST purposes
            "enddatetime": f'{enddate}T07:00-0000', 
            "market_run_id": market_run_id,
            "version": version,  
            "node": node,
            }

        response = requests.get(url, params=params)
        try: # reading the file
            with ZipFile(BytesIO(response.content)) as z:
                for filename in z.namelist():
                    with z.open(filename) as f:
                        df = pd.read_csv(f) 
                        df.to_csv(f'pull#{counter}.csv') # creating chunk
                        if '<?xml version="1.0" encoding="UTF-8"?>' in df.columns:
                            pass
                        else: # only appending files that pull- error handling 
                            files.append(f'pull#{counter}.csv') # creating a list of all files I need to combine
        except Exception as e: # if there are errors- usually doesn't come to this. can probably delete
            print(f"Error message: {e}")

    # cleaning function!
    def cleanFile(filename):
        df = pd.read_csv(filename) # reading in file
        all_drop = ['NODE_ID_XML', 'NODE_ID', 'PNODE_RESMRID', 'OPR_DT', 'OPR_HR', 'OPR_INTERVAL', 'XML_DATA_ITEM', 'POS', 'GROUP', 'GRP_TYPE', 'MARKET_RUN_ID', 'Unnamed: 0', 'INTERVAL_START_TIME'] # dropping unneeded columns
        valid_columns = [col for col in all_drop if col in df.columns]
        df = df.drop(columns=valid_columns) # dropping conditionally- if they are in the df, they are dropped. prevents errors. 
        df = df.sort_values(['INTERVALSTARTTIME_GMT']) # sorting by time/date

        # changing timezone to MST
        df['INTERVALSTARTTIME_GMT'] = df['INTERVALSTARTTIME_GMT'].str.replace('-00:00', '').str.replace('T', ' ') # getting rid of seconds
        df['INTERVALENDTIME_GMT'] = df['INTERVALENDTIME_GMT'].str.replace('-00:00','').str.replace('T',' ') # getting rid of seconds
        df['INTERVALSTARTTIME_GMT'] = pd.to_datetime(df['INTERVALSTARTTIME_GMT']) # changing to datetime
        df['INTERVALSTARTTIME_GMT'] = df['INTERVALSTARTTIME_GMT'] - timedelta(hours=7) # shifting hours
        df['INTERVALENDTIME_GMT'] = pd.to_datetime(df['INTERVALENDTIME_GMT']) # changing to datetime
        df['INTERVALENDTIME_GMT'] = df['INTERVALENDTIME_GMT'] - timedelta(hours=7) # shifting hours
        df.rename(columns={'INTERVALSTARTTIME_GMT': 'INTERVALSTARTTIME_MST', 'INTERVALENDTIME_GMT': 'INTERVALENDTIME_MST'}, inplace=True)
        
        # conditional cleaning
        if 'LMP_TYPE' in df.columns: # making LMP_TYPE more readable
                df['LMP_TYPE'] = df['LMP_TYPE'].replace({'LMP': 'LMP', 'MCC': 'Congestion', 'MCE':'Energy', 'MCL': 'Loss', 'MGHG': 'Greenhouse Gas'})
        if 'MW' in df.columns: # rounding DAM and HASP
            df['MW'] = df['MW'].round(4)
        if 'VALUE' in df.columns:# rounding RTM
            df['VALUE'] = df['VALUE'].round(4)
        if 'PRC' in df.columns:# rounding FMM
            df['PRC'] = df['PRC'].round(4)

        # splitting date into smaller columns for readability and grouping
        df['INTERVALSTARTTIME_MST'] = df['INTERVALSTARTTIME_MST'].astype(str) # changing start time from date to string so I can split it
        df[['Year', 'Month','Day']] = df['INTERVALSTARTTIME_MST'].str.split('-',expand=True)
        df[['Day', 'Time']] = df['Day'].str.split(' ',expand=True)
        df[['Hour (MST)','Minute', 'Seconds']] = df['Time'].str.split(':',expand=True)
        df = df.drop(columns=['Time', 'Seconds']) # ditching time and seconds
        df.to_csv(filename) # replacing file with cleaned version

    # missing values function
    def fill_missing_values(filename):
        df = pd.read_excel(filename)
        # finding and creating rows for missing intervals
        if market_run_id == 'DAM':
            df['INTERVALSTARTTIME_MST'] = pd.to_datetime(df['INTERVALSTARTTIME_MST']).dt.floor('60min') # makes it some sort of set so I can subtact it later
            full_range = pd.date_range(start=df['INTERVALSTARTTIME_MST'].min(), end=df['INTERVALSTARTTIME_MST'].max(), freq='60min') # finding all intervals I should have
        elif market_run_id == 'RTM':
            df['INTERVALSTARTTIME_MST'] = pd.to_datetime(df['INTERVALSTARTTIME_MST']).dt.floor('5min') # makes it some sort of set so I can subtact it later
            full_range = pd.date_range(start=df['INTERVALSTARTTIME_MST'].min(), end=df['INTERVALSTARTTIME_MST'].max(), freq='5min') # finding all intervals I should have
        else: # HASP AND FFM
            df['INTERVALSTARTTIME_MST'] = pd.to_datetime(df['INTERVALSTARTTIME_MST']).dt.floor('15min') # makes it some sort of set so I can subtact it later
            full_range = pd.date_range(start=df['INTERVALSTARTTIME_MST'].min(), end=df['INTERVALSTARTTIME_MST'].max(), freq='15min') # finding all intervals I should have
        full_df = pd.DataFrame({'INTERVALSTARTTIME_MST': full_range}) # df for all intervals I need
        if 'Greenhouse Gas' in full_df.columns: # conditional logic for HASP
            result = full_df.merge( # combining data with full intervals, putting in nulls for vals in rows that had missing intervals 
                df[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Hour (MST)', 'Minute','Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']],  # Include original data columns
                on='INTERVALSTARTTIME_MST',
                how='outer'
            )
        else:
            result = full_df.merge( # combining data with full intervals, putting in nulls for vals in rows that had missing intervals 
                df[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Hour (MST)', 'Minute','Congestion', 'Energy', 'Loss', 'LMP']],  # Include original data columns
                on='INTERVALSTARTTIME_MST',
                how='outer'
            )
        # filling in missing values
        if market_run_id == 'DAM':
            interval_end = result['INTERVALSTARTTIME_MST'] + pd.Timedelta(minutes=60) # finding starttime and adding 15 minutes
        elif market_run_id == 'RTM':
            interval_end = result['INTERVALSTARTTIME_MST'] + pd.Timedelta(minutes=5) # finding starttime and adding 15 minutes
        else: # ffm and hasp
            interval_end = result['INTERVALSTARTTIME_MST'] + pd.Timedelta(minutes=15) # finding starttime and adding 15 minutes
        result['INTERVALENDTIME_MST']=result['INTERVALENDTIME_MST'].fillna(interval_end) # filling in empty values
        result = result.sort_values(['INTERVALSTARTTIME_MST', 'NODE']) # making sure it is sorted by node so I can backfill
        result['NODE'] = result['NODE'].bfill() # backfilling node
        result['INTERVALSTARTTIME_MST'] = result['INTERVALSTARTTIME_MST'].astype(str) # changing start time from date to string so I can split it
        for row in result: # filtering through all rows
            if result['INTERVALENDTIME_MST'].isnull: # if it has missing values, fill them in!
                result[['Year', 'Month','Day']] = result['INTERVALSTARTTIME_MST'].str.split('-',expand=True) # splitting interval to make yr, mnth, hr, etc
                result[['Day', 'Time']] = result['Day'].str.split(' ',expand=True)
                result[['Hour (MST)','Minute', 'Seconds']] = result['Time'].str.split(':',expand=True)
        result = result.drop(columns=['Time', 'Seconds'])
        # filling in LMP values
        result = result.sort_values(['Hour (MST)', 'Minute'])
        if 'Greenhouse Gas' in result.columns: # conditional logic for HASP
            LMP_list = ['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']    
        else:
            LMP_list = ['Congestion', 'Energy', 'Loss', 'LMP']
        for column in LMP_list: # making a for loop for efficiency purposes    
            result[column] = result[column].bfill() # can backfill here because of the new sorting- takes previous day's same hour and minute intvl price.
        result = result.sort_values('INTERVALSTARTTIME_MST') # setting to original sorting
        # reordering columns 
        if 'Greenhouse Gas' in result.columns: 
            result = result[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute','LMP','Congestion', 'Energy', 'Loss', 'Greenhouse Gas' ]]
        else:
            result = result[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute', 'LMP','Congestion', 'Energy', 'Loss' ]]
        result.to_excel(filename) # should replace file with the filled version

    # monthly average sheet function (12x24 info)
    def monthly_average(filename):
        df = pd.read_excel(filename)
        if 'Greenhouse Gas' in df.columns: # doesn't have greenhouse gas
            df_avg = df.groupby(['NODE', 'Year', 'Month'], as_index=False)[['Congestion', 'Energy', 'Greenhouse Gas', 'LMP', 'Loss']].mean() # grouping
            df_avg[['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']] = df_avg[['Congestion', 'Energy', 'Loss', 'Greenhouse Gas' 'LMP']].round(4)
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Congestion', 'Energy', 'Greenhouse Gas', 'Loss', 'LMP']] # reordering column names
        else:
            df_avg = df.groupby(['NODE', 'Year', 'Month'], as_index=False)[['Congestion', 'Energy', 'LMP', 'Loss']].mean() # grouping
            df_avg[['Congestion', 'Energy', 'Loss','LMP']] = df_avg[['Congestion', 'Energy', 'Loss','LMP']].round(4)
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Congestion', 'Energy', 'Loss', 'LMP']] # reordering column names
        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx', engine='openpyxl', mode='a') as writer: # adding sheet to excel file
            df_avg.to_excel(writer, sheet_name='Monthly Average', index=False)       
        plt.plot(df_avg['Month'], df_avg['LMP']) # creating monthly average line chart
        plt.title(f'Monthly Average LMP {node}')
        plt.ylabel('Avg $/MWH')
        plt.xlabel('Month')
        plt.savefig(f'{output_file_path}/monthlyline.png') # saving png
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Monthly Average']
        img_path = f'{output_file_path}/monthlyline.png'
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(img_path)
        sheet.add_image(img, 'I2') # adding png to sheet
        wb.save(filename) # saving workbook with added pngs
        plt.close() # closing plt so it doesn't combine with other chart later
        os.remove(img_path) # removing png from folder

    # hourly average sheet function
    def hourly_average(filename): # creating a new sheet in the same excel file for hourly averages
        df = pd.read_excel(filename)
        if 'Greenhouse Gas' in df.columns: # doesn't have greenhouse gas
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Day', 'Hour (MST)'], as_index=False)[['Congestion', 'Energy', 'Greenhouse Gas', 'LMP', 'Loss']].mean().round(4) # grouping
            df_avg[['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']] = df_avg[['Congestion', 'Energy', 'Loss', 'Greenhouse Gas' 'LMP']].round(4)            
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Congestion', 'Energy', 'Greenhouse Gas', 'Loss', 'LMP']] # reordering column names
        else:
            df_avg = df.groupby(['NODE', 'Year', 'Month', 'Day', 'Hour (MST)'], as_index=False)[['Congestion', 'Energy', 'LMP', 'Loss']].mean().round(4) # grouping
            df_avg[['Congestion', 'Energy', 'Loss','LMP']] = df_avg[['Congestion', 'Energy', 'Loss','LMP']].round(4)
            df_avg = df_avg[['NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Congestion', 'Energy', 'Loss', 'LMP']] # reordering column names
        
        count = (df_avg['LMP'] < 0).sum() # counting how many hours LMP is below 0

        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # writing hourly averages sheet
            df_avg.to_excel(writer, sheet_name='Hourly Average', index=False) # adding sheet to excel file
            # writing summary statistics sheet, using the hourly averages df
            row = 13
            pd.DataFrame([['Number of hours LMP is below 0:']]).to_excel(writer, sheet_name = 'Summary Statistics', startrow = row, header=False, index=False)
            row += 1
            pd.DataFrame([[count]]).to_excel(writer, sheet_name = 'Summary Statistics', startrow = row, header=False, index=False)
            row += 2 # adding a blank row in between
            pd.DataFrame([['Duration Curve']]).to_excel(writer, sheet_name = 'Summary Statistics', startrow = row, header=False, index=False)
            pd.DataFrame([['Lowest 5% Zoom']]).to_excel(writer, sheet_name = 'Summary Statistics', startrow = row, startcol=12, header=False, index=False)
            row += 1


        # heat map!
        pivot = df.pivot_table(index='Hour (MST)', columns='Month', values='LMP', aggfunc='mean') # creating a new table section df
        plt.figure(figsize=(10,8))
        sns.heatmap(pivot, annot=True, cmap='RdYlGn_r', fmt='.0f', cbar_kws={'label':'$/MWH'}) # creating heatmap
        plt.title(f'12x24 Heatmap {node}')
        plt.savefig(f'{output_file_path}/heatmap.png') # saving png
        plt.close() # closing plt so it doesn't combine with other chart later        
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Hourly Average']
        img_path = f'{output_file_path}/heatmap.png'
        from openpyxl.drawing.image import Image as XLImage # opening excel file
        img = XLImage(img_path)
        sheet.add_image(img, 'L3') # adding chart image to sheet
        wb.save(filename)
        os.remove(img_path) # removing png from folder

    # summary statistics sheet function
    def summary_statistics(filename): 
        df = pd.read_excel(filename)
        if 'Greenhouse Gas' in df.columns: # conditional logic for finding applicable column
            cols = ['Congestion', 'Energy', 'Loss', 'Greenhouse Gas', 'LMP']
        else:
            cols = ['Congestion', 'Energy', 'Loss', 'LMP']
        desc=df[cols].describe() # finding what I want to display

        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer: # adding sheet to excel file
            # summary statistics for 4 elements of LMP
            row = 0
            pd.DataFrame([['Summary Statistics']]).to_excel(writer, sheet_name = 'Summary Statistics', startrow = row, header=False, index=False) # title
            row += 1
            desc.to_excel(writer, sheet_name = 'Summary Statistics', startrow=row, header=True, index=True) # writing summary statistics to df
            row += 13 # adding space in between  


        # maybe include a macro for readability formatting

    # duration chart function
    def duration_chart(filename): 
        # cleaning to get chart columns
        df = pd.read_excel(filename)
        df = df.sort_values('LMP', ascending=False) # sorting by price/LMP high to low
        duration_counts = df['LMP'].value_counts() # counting how many there are of each LMP value
        total_count = df['LMP'].value_counts().sum() # counting how many values total
        df['duration_count'] = df['LMP'].map(duration_counts) # this represents how many there are of that specific value
        df['percent'] = df['duration_count']/total_count # percentage value
        chart_lmp = df[['LMP', 'percent']].drop_duplicates().copy() # making sure we aren't counting percents more than once (if a LMP value exists multiple times)
        chart_lmp['xval'] = chart_lmp['percent'].cumsum() # cumulative sum adds the rows value to all the previous rows values
        xval_map = dict(zip(chart_lmp['LMP'], chart_lmp['xval'])) # creating a map of new percentages?
        df['xval'] = df['LMP'].map(xval_map) # adding to OG DF
        df = df[['LMP', 'xval']] # limiting df to only the necessary columns

        # writting to a sheet and hiding it 
        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name='Hidden Duration Chart Data', index=False) # writing data to new sheet
        
        wb = openpyxl.load_workbook(filename) # finding workbook with filename
        ws = wb['Hidden Duration Chart Data'] # referencing
        ws.sheet_state = 'hidden' # hiding the sheet

        sheet = wb['Summary Statistics'] # opening summary stats sheet
        sheet['A1'].font = Font(bold=True) # bolding titles
        sheet['A14'].font = Font(bold=True)
        sheet['A17'].font = Font(bold=True)      
        sheet['M17'].font = Font(bold=True)
        plt.figure()
        plt.scatter(df['xval'], df['LMP'], s=3) # creating duration chart using a 
        plt.axhline(y=0, color ='black') # creating 0 axis line
        plt.title(f'Duration Chart {node}') 
        plt.ylabel('$/MWh')
        plt.xlabel('% of Time')
        plt.grid()
        plt.savefig(f'{output_file_path}/durationchart.png') # creating an image of the chart
        plt.close()
        img_path2 = f'{output_file_path}/durationchart.png' # making image path
        from openpyxl.drawing.image import Image as XLImage # writing to excel
        img2 = XLImage(img_path2)
        sheet.add_image(img2, 'B19') # adding chart image to sheet

        # lowest 5% zoom
        last5index = int(0.95*len(df)) # finding last 5% of data
        plt.figure()
        plt.scatter(df['xval'], df['LMP'], s=3)
        x_min = df['xval'].iloc[last5index]  # First x-value in last 5%
        x_max = df['xval'].iloc[-1]  # Last x-value
        plt.xlim(x_min,x_max) # creating the zoomed x axis
        plt.axhline(y=0, color='black') # adding in 0 line for reference
        plt.title(f'Lowest 5% Zoom {node}')
        plt.ylabel('$/MWh')
        plt.xlabel('% of Time')
        plt.grid()
        plt.savefig(f'{output_file_path}/lowest5zoom.png') # saving png
        plt.close()
        img_path = f'{output_file_path}/lowest5zoom.png'
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(img_path)
        sheet.add_image(img, 'L19') # adding png to sheet
        wb.save(filename) # saving workbook to original file name
        os.remove(img_path2) # removing pngs from folder space
        os.remove(img_path)


    # user inputs map
    map = { # market name, market_run_id, query_name, version - needed for parameters
        ('DAM'): ('DAM', 'PRC_LMP', 1), 
        ('RTM'): ('RTM', 'PRC_INTVL_LMP', 3),
        ('HASP'): ('HASP', 'PRC_HASP_LMP', 1),
        ('FFM'): ('RTPD', 'PRC_RTPD_LMP', 2,), 
    }
    
    market_run_id, queryname, version = map.get((market_run_id), ('unknown')) # getting variables from map based on user input
    startdate = datetime.strptime(startdate, '%m/%d/%y').date() # formats it to work with datetime package
    enddate = datetime.strptime(enddate, '%m/%d/%y').date()
    difference = enddate - startdate # counts days in between
    days = difference.days # making a counter for my loop bc .days is readonly

    # logic for using the API, cleaning, and formatting as an xlsx file.
    files = [] # creating an empty list for my csv files 
    df_list = [] # empty list that stores the actual rows of the files later (for concatenation)
    counter = 0 # initializing, used to label different files
    timestamp = datetime.now() # using for filename
    timestamp = timestamp.strftime('%m-%d-%Y %H%M')

    if difference.days > 30: # logic that breaks the whole date range into chunks of 30 days to comply with the API
        while days > 0: # until we reach the end date
            if days < 30: # when we get below 30 days left
                print(f'Pulling data for {startdate} to {enddate}.') 
                pull_request(startdate, enddate)
                break # ending loop once I get to the end date
            nextdate = startdate + timedelta(days=30) # creating a chunk
            print(f'Pulling data for {startdate} to {nextdate}.') 
            pull_request(startdate, nextdate)
            days -= 30 # updating this counter
            counter += 1 
            startdate = nextdate # update start date to be next date
            time.sleep(10) # avoiding query limits
        for file in files: # if tabbed over, it cleans each pull before combining, but is bad at error handling
            cleanFile(file) 
            df_list.append(pd.read_csv(file))
    
        # combining files and cleaning them up a little
        df_combined = pd.concat(df_list, ignore_index=True) # combining 30 day chunks
        cond_drop = ['Unnamed: 0.1', 'Unnamed: 0'] # dropping these if they exist
        conditional_drop = [col for col in cond_drop if col in df_combined.columns]
        df_combined = df_combined.drop(columns=conditional_drop)
        df_combined = df_combined.drop_duplicates() # dropping duplicate 

        # pivoting table and reordering columns (for first sheet)
        if 'MW' in df_combined.columns: # this one will do nothing for the DAM pull, except maybe make a duplicate page
            df_combined = pd.pivot_table(df_combined, values='MW', index=['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute'], columns='LMP_TYPE') # breaking out LMP_TYPE columns, keeping the other indexed columns
            df_combined = df_combined.reset_index() # resetting index to work with column names
            if 'Greenhouse Gas' in df_combined.columns: # DAM 
                df_combined = df_combined[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute', 'Congestion', 'Energy', 'Greenhouse Gas', 'Loss', 'LMP']]
            else: # HASP 
                df_combined = df_combined[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute', 'Congestion', 'Energy', 'Loss', 'LMP']]
        if 'VALUE' in df_combined.columns: # RTM , I think
            df_combined = pd.pivot_table(df_combined, values='VALUE', index=['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute'], columns='LMP_TYPE')
            df_combined = df_combined.reset_index()
            df_combined = df_combined[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute', 'Congestion', 'Energy', 'Greenhouse Gas', 'Loss', 'LMP']]
        if 'PRC' in df_combined.columns: # FFM , I think
            df_combined = pd.pivot_table(df_combined, values='PRC', index=['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute'], columns='LMP_TYPE')
            df_combined = df_combined.reset_index()
            df_combined = df_combined[['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute', 'Congestion', 'Energy', 'Greenhouse Gas', 'Loss', 'LMP']]
        
        # pushing to excel file, deleting csv chunks
        # os.makedirs(output_file_path, exist_ok=True) # making sure the directory exists?
        file = f'{output_file_path}/{market_run_id} {timestamp}.xlsx'
        with pd.ExcelWriter(file, engine='openpyxl') as writer: 
            df_combined.to_excel(writer, sheet_name = 'Report', index=False) # writing initial report to an xlsx file
        fill_missing_values(file)        
        monthly_average(file) # writing and adding monthly average sheet to file        
        hourly_average(file) # writing and adding hourly average sheet to file
        summary_statistics(file)
        duration_chart(file)
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet1']
        sheet.title='Report'
        wb.save(file)

        # deleting all of the csv 30 day chunk files from the folder
        if getattr(sys, 'frozen', False): # finding file path to wherever GUI is stored
            application_path = os.path.dirname(sys.executable)
        for csv in files: # deleting extra files
            filepath = os.path.join(application_path, csv) # creating path to CSV file that will populate in the GUI folder
            os.remove(filepath) # deleting files

        status_lbl.configure(text='Finished!') # updating status label 
        root.update()            

    if difference.days <= 30: # logic for user requests that are less than 30 days (doesn't need chunking)
        print(f'Pulling data for {startdate} to {enddate}.')
        pull_request(startdate, enddate)
        cleanFile('pull#0.csv')
        df = pd.read_csv('pull#0.csv')
        if 'MW' in df.columns: # this one will do nothing for the DAM pull
            df = pd.pivot_table(df, values='MW', index=['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute'], columns='LMP_TYPE') # breaking out LMP_TYPE columns, keeping the other indexed columns
        if 'VALUE' in df.columns:
            df = pd.pivot_table(df, values='VALUE', index=['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute'], columns='LMP_TYPE')
        if 'PRC' in df.columns:
            df = pd.pivot_table(df, values='PRC', index=['INTERVALSTARTTIME_MST', 'INTERVALENDTIME_MST', 'NODE', 'Year', 'Month', 'Day', 'Hour (MST)', 'Minute'], columns='LMP_TYPE')
        df = df.reset_index()
        with pd.ExcelWriter(f'{output_file_path}/{market_run_id} {timestamp}.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name = 'Report', index=False) # writing initial report to an xlsx file
        file = f'{output_file_path}/{market_run_id} {timestamp}.xlsx'
        fill_missing_values(file)
        monthly_average(file) # writing and adding monthly average sheet to file        
        hourly_average(file) # writing and adding hourly average sheet to file
        summary_statistics(file)
        duration_chart(file)
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet1']
        sheet.title='Report'
        wb.save(file)

        # deleting csv from folder
        if getattr(sys, 'frozen', False): # finding file path to wherever GUI is stored
            application_path = os.path.dirname(sys.executable)
        filepath = os.path.join(application_path, csv) # creating path to CSV file that will populate in the GUI folder
        os.remove(filepath) # deleting files

        status_lbl.configure(text='Finished!')
        root.update() # updating status label

   
# button/widget functions
def submit(): # after user gives all inputs, runs all of the backend code
    status_lbl.configure(text='Running...')
    root.update() # updating status label
    market_run_id=MRIDDropdown.get() # grabbing market_run_id based on user input
    backend(market_run_id, startdate, enddate) # calling backend code

def findStartDate(): # for selecting the start date
    global startdate
    startdate = cal.get_date()
    startdate_label.configure(text=f'Start date: {startdate}')

def findEndDate(): # for selecting the end date
    global enddate
    enddate = cal.get_date()
    enddate_label.configure(text=f'End date: {enddate}')

def select_output_file(): # for selecting excel output file path
    global output_file_path
    directory = filedialog.askdirectory(title='Select output directory')
    if directory:
        output_file_path = directory
        output_file_label.configure(text=directory) # displaying file path
    else:
        output_file_label.configure(text='No directory selected yet') # if they tried to submit without a filepath

def update_report_lbl(choice): # displays the report name, based on the user chosen market type
    if choice == 'DAM':
        reportname = 'Locational Marginal Prices'
    if choice == 'RTM':
        reportname = 'Interval Locational Marginal Prices'
    if choice == 'HASP':
        reportname = 'Hour Ahead Locational Marginal Prices'
    if choice == 'FFM':
        reportname = 'FMM Locational Marginal Prices'
    report_lbl.configure(text=f'{reportname}')
    root.update()


# tkinter program
root = CTk() # initializing window
root.geometry('800x600') # setting size
set_appearance_mode('light') # can also be light

node_var=tk.StringVar()

startdate = None # initializing
enddate = None

# widgets
MRID_label = CTkLabel(root, text = 'Market Type:', font=('Arial',15), text_color='#04033A')
MRIDDropdown = CTkComboBox(master=root, values=['DAM', 'RTM', 'HASP', 'FFM'], command=update_report_lbl)

report_lbl = CTkLabel(root, text = 'Locational Marginal Prices', font=('Arial',15), text_color='#04033A')

cal = Calendar(root, selectmode ='day',
            year=2024, month =1, # defaults
            day = 1, font=('Arial', 15))

chooseStartDate = CTkButton(root, text='Choose Start Date', command=findStartDate, corner_radius=26,fg_color='#162157', hover_color='#6D7DCF')
chooseEndDate = CTkButton(root, text='Choose End Date', command=findEndDate, corner_radius=26,fg_color='#162157', hover_color='#6D7DCF')

startdate_label = CTkLabel(root, text= 'Start Date: ', font=('Arial',15), text_color='#04033A') 
enddate_label = CTkLabel(root, text='End Date: ', font=('Arial',15), text_color='#04033A')

node_label = CTkLabel(root, text='Node(s):', font=('Arial',15), text_color='#04033A')
node_entry = CTkEntry(root, textvariable = node_var, font=('Arial',15), text_color='#04033A')

sub_btn=CTkButton(master=root,text = 'Submit', command = submit, corner_radius=32,fg_color='#162157', hover_color='#6D7DCF') 

output_file_button = CTkButton(root, text='Select Output File Path', command=select_output_file, corner_radius=32,fg_color='#162157', hover_color='#6D7DCF')
output_file_label = CTkLabel(root, text='No path selected', font=('Arial',10), text_color='#04033A')

status_lbl = CTkLabel(root, text='', font=('Arial',15), text_color='#04033A')

title_lbl = CTkLabel(root, text='CAISO OASIS DATA', font=('Arial',20, 'bold'), text_color='#04033A')

# grid- where all the widgets are displayed on the GUI
cal.grid(row=6,column=0)
chooseStartDate.grid(row=4,column=0) 
chooseEndDate.grid(row=5,column=0)
MRID_label.grid(row=1, column=1)
MRIDDropdown.grid(row=1, column=2)
node_label.grid(row=3, column=1)
node_entry.grid(row=3,column=2)
sub_btn.grid(row=6,column=2)
startdate_label.grid(row=4, column=1)
enddate_label.grid(row=5, column=1)
output_file_button.grid(row=1, column=0)
output_file_label.grid(row=2, column=0)
title_lbl.grid(row=0, column=2)
report_lbl.grid(row=2,column=2)
status_lbl.grid(row=7,column=2) 

root.mainloop() # performing an infinite loop for the window to display